"""去除 Excel 重複商品（依分潤連結去重），重新編號"""
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = os.getenv('KEYWORD_OUTPUT', r'D:\Users\user\Desktop\蝦皮影片專案\蝦皮選品_2026年5月.xlsx')

wb = openpyxl.load_workbook(OUTPUT)
ws = wb.active

rows = list(ws.iter_rows(min_row=2, values_only=True))
print(f'原始筆數：{len(rows)}')

seen_links = set()
unique = []
for row in rows:
    link = str(row[2] or '').strip()  # 分潤連結在第3欄（index 2）
    if not link or link in seen_links:
        continue
    seen_links.add(link)
    unique.append(row)

print(f'去重後筆數：{len(unique)}')

# 清除舊資料（保留 header）
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.value = None

# 重新寫入
thin = Side(style='thin', color='DDDDDD')
bd = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, row in enumerate(unique):
    ri = i + 2
    row_data = list(row)
    row_data[0] = i + 1  # 重新編號
    fill = PatternFill('solid', fgColor='FFF5F5' if ri % 2 == 0 else 'FFFFFF')
    for ci, val in enumerate(row_data, 1):
        c = ws.cell(ri, ci, val)
        c.fill = fill
        c.border = bd
        c.font = Font(name='Arial', size=10, color='0563C1', underline='single') if ci == 3 \
                 else Font(name='微軟正黑體', size=10)
        c.alignment = Alignment(horizontal='left' if ci in (2, 3) else 'center', vertical='center')
    ws.row_dimensions[ri].height = 32

# 清除多餘舊行
for ri in range(len(unique) + 2, ws.max_row + 1):
    for ci in range(1, 11):
        ws.cell(ri, ci).value = None

wb.save(OUTPUT)
print(f'完成！已儲存：{OUTPUT}')
