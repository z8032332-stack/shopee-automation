# -*- coding: utf-8 -*-
"""
讀取現有 Excel，對每筆商品查影片數量，
保留影片數 >= MIN_VIDEO_COUNT 的前 KEEP_TOP 筆，覆寫回原檔。
"""
import sys, io, asyncio, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))

from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT         = os.getenv('KEYWORD_OUTPUT', r'D:\Users\user\Desktop\蝦皮影片專案\蝦皮關鍵字選品_2026年4-5月.xlsx')
MIN_VIDEO_COUNT = 3
KEEP_TOP        = 50

# 從分潤連結解析 shop_id / item_id
def parse_ids(link):
    # 格式1: /product/{shopid}/{itemid}
    m = re.search(r'/product/(\d+)/(\d+)', link or '')
    if m: return m.group(1), m.group(2)
    # 格式2: i.{shopid}.{itemid}
    m = re.search(r'i\.(\d+)\.(\d+)', link or '')
    if m: return m.group(1), m.group(2)
    return '', ''

async def safe_eval(page, js, retries=3):
    for i in range(retries):
        try: return await page.evaluate(js)
        except:
            if i < retries - 1: await asyncio.sleep(1.5)
    return None

async def get_video_count(page, sid, iid, debug=False):
    if not sid or not iid:
        if debug: print(f'    [debug] 空 ID: sid={sid!r} iid={iid!r}')
        return 0
    js = """
    (async () => {
      try {
        const r = await fetch('https://shopee.tw/api/v4/item/get?itemid=IID&shopid=SID', {credentials:'include'});
        const d = await r.json();
        if (!d || !d.data) return {err: 'no data', keys: Object.keys(d||{})};
        return (d.data.video_info_list) ? d.data.video_info_list.length : 0;
      } catch(e) { return {err: e.toString()}; }
    })()
    """.replace('IID', str(iid)).replace('SID', str(sid))
    result = await safe_eval(page, js)
    if debug: print(f'    [debug] sid={sid} iid={iid} -> {result}')
    if isinstance(result, int): return result
    return 0

async def main():
    print('=' * 50)
    print(f'篩選影片數 >= {MIN_VIDEO_COUNT}，保留前 {KEEP_TOP} 筆')
    print('=' * 50)

    # 讀 Excel
    wb = openpyxl.load_workbook(OUTPUT)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f'原始筆數：{len(rows)}')

    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp('http://127.0.0.1:9222')
        ctx = browser.contexts[0]
        pg = next((pg for pg in ctx.pages
                   if pg.url.startswith('https://shopee.tw') and 'affiliate' not in pg.url), None)
        if not pg:
            pg = await ctx.new_page()
        # 若停在驗證碼頁，重新導向首頁
        if 'captcha' in pg.url or 'verify' in pg.url or pg.url.strip('/') == 'https://shopee.tw':
            print('  重新導向 shopee.tw 首頁...')
            await pg.goto('https://shopee.tw', wait_until='domcontentloaded', timeout=20000)
            await asyncio.sleep(3)
        print(f'shopee 頁面：{pg.url[:60]}')

        results = []
        for i, row in enumerate(rows):
            link = row[2] if len(row) > 2 else ''   # 第3欄 = 分潤連結
            name = row[1] if len(row) > 1 else ''
            sid, iid = parse_ids(link)
            debug = (i < 2)  # 只 debug 前2筆
            if debug: print(f'    [debug] link={str(link)[:80]}')
            cnt = await get_video_count(pg, sid, iid, debug=debug)
            status = f'✅ {cnt}支' if cnt >= MIN_VIDEO_COUNT else f'❌ {cnt}支'
            print(f'  [{i+1:2}/{len(rows)}] {str(name)[:25]:25s} {status}')
            if cnt >= MIN_VIDEO_COUNT:
                results.append(row)
            if len(results) >= KEEP_TOP:
                print(f'  已達 {KEEP_TOP} 筆，停止檢查')
                break
            await asyncio.sleep(0.3)

    print(f'\n符合條件：{len(results)} 筆')

    if len(results) == 0:
        print('⚠️  0 筆符合條件，可能是 shopee.tw 驗證碼擋住，不覆寫 Excel，請先手動過驗證碼再重跑。')
        return

    # 重寫 Excel（只保留符合的）
    thin = Side(style='thin', color='DDDDDD')
    bd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdrs = ['編號','品名','分潤連結','價格','分潤率','銷量','對應關鍵字','文案','標題','狀態']

    # 清掉舊資料列（保留 header）
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # 寫回
    for i, row in enumerate(results):
        ri = i + 2
        fill = PatternFill('solid', fgColor='FFF5F5' if ri % 2 == 0 else 'FFFFFF')
        vals = list(row)
        vals[0] = i + 1   # 重編號
        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci, val)
            c.fill = fill; c.border = bd
            c.font = Font(name='Arial', size=10, color='0563C1', underline='single') if ci == 3 \
                     else Font(name='微軟正黑體', size=10)
            c.alignment = Alignment(horizontal='left' if ci in (2, 3) else 'center',
                                    vertical='center', wrap_text=(ci == 2))
        ws.row_dimensions[ri].height = 32

    ws.auto_filter.ref = f'A1:J{len(results)+1}'
    wb.save(OUTPUT)
    print(f'Excel 已覆寫：{OUTPUT}（共 {len(results)} 筆）')

asyncio.run(main())
