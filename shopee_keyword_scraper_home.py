# [HOME/COMPANY VERSION] 家用/公司通用版 - 2026年4-5月
# 與公司版差異：直接 CDP 連線至 Chrome port 9222，不使用 profile 登入
# OUTPUT 路徑從 .env 讀取（KEYWORD_OUTPUT）
import sys, io, asyncio, json, re, os, random
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))

from playwright.async_api import async_playwright
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ALL_KEYWORDS = [
    # 3C / 電子（影片比例最高）
    '掃地機器人','無線吸塵器','空氣清淨機','除濕機',
    '智能手錶','運動手錶','藍芽喇叭','無線耳機',
    '行動電源','電動牙刷','洗臉機','美顏儀',
    # 廚房家電（幾乎都有 demo 影片）
    '氣炸鍋','咖啡機','果汁機','電熱水壺','烤箱',
    '全自動奶泡機','電動開罐器','食物調理機',
    # 美妝儀器（demo 影片非常普遍）
    '射頻美容儀','微電流美容儀','電動睫毛夾',
    '美髮梳','負離子吹風機','捲髮棒','直捲兩用夾',
    # 按摩/健康
    '按摩槍','筋膜槍','眼部按摩儀','頸部按摩儀',
    '電動足浴機','背部按摩器',
    # 夏季家電
    '電風扇','循環扇','行動冷氣','水冷扇',
]
BLACKLIST        = ['藥','酒','菸','棉花棒','化妝棉']
TARGET           = int(os.getenv('KEYWORD_TARGET',        '50'))
MIN_SALES        = int(os.getenv('KEYWORD_MIN_SALES',     '500'))
MAX_PER_KW       = int(os.getenv('KEYWORD_MAX_PER_KW',    '5'))
MIN_VIDEO_REVIEW = int(os.getenv('KEYWORD_MIN_VID_REVIEW','3'))   # 評論區影片數下限
REVIEW_DAYS      = int(os.getenv('KEYWORD_REVIEW_DAYS',  '30'))   # 最新評論幾天內
BATCH            = TARGET * 3
OUTPUT           = os.getenv('KEYWORD_OUTPUT', r'D:\Users\user\Desktop\蝦皮影片專案\蝦皮關鍵字選品_2026年4-5月.xlsx')

# ── 商品去重：每次跑完記錄已選商品 ID，下次自動排除 ──
PRODUCT_HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'product_history.json')

def load_product_history():
    """載入歷史商品 ID 集合 {uid: date}"""
    if os.path.exists(PRODUCT_HISTORY_FILE):
        with open(PRODUCT_HISTORY_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_product_history(new_products):
    """把本次選到的商品 uid 存進歷史"""
    from datetime import date
    h = load_product_history()
    today = date.today().isoformat()
    for p in new_products:
        uid = f"{p['shop_id']}_{p['item_id']}"
        h[uid] = today
    with open(PRODUCT_HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(h, f, ensure_ascii=False, indent=2)
    print(f'  商品歷史已更新（累計 {len(h)} 筆）：{PRODUCT_HISTORY_FILE}')

KEYWORDS = ALL_KEYWORDS[:]
random.shuffle(KEYWORDS)  # 關鍵字每次隨機順序

def is_bl(name): return any(k in name for k in BLACKLIST)
def pp(raw):
    try:
        v = int(raw)
        return str(v // 100000) if v > 100000 else str(v)
    except: return str(raw)

async def safe_eval(page, js, retries=3):
    for i in range(retries):
        try: return await page.evaluate(js)
        except:
            if i < retries - 1: await asyncio.sleep(1.5)
    return None

async def search_kw(page, kw, limit=20):
    enc = quote(kw)
    js = """
    (async () => {
      try {
        const r = await fetch('https://affiliate.shopee.tw/api/v3/offer/product/list?list_type=0&sort_type=1&page_offset=0&page_limit=LIMIT&keyword=KW', {credentials:'include'});
        const d = await r.json();
        return d;
      } catch(e) { return {error: e.toString()}; }
    })()
    """.replace('LIMIT', str(limit)).replace('KW', enc)
    d = await safe_eval(page, js) or {}
    items = []
    for it in (d.get('data') or {}).get('list') or []:
        info = it.get('batch_item_for_item_card_full') or {}
        name = info.get('name', '')
        if is_bl(name): continue
        sales = info.get('sold') or info.get('historical_sold') or 0
        if sales < MIN_SALES: continue
        price = pp(info.get('price') or info.get('price_min', 0))
        shop_id = str(info.get('shopid', ''))
        item_id = str(info.get('itemid', '') or it.get('item_id', ''))
        items.append({
            'name': name, 'price': price, 'sales': sales, 'keyword': kw,
            'comm_rate': it.get('default_commission_rate', '') or it.get('seller_commission_rate', ''),
            'affiliate_link': it.get('long_link', ''),
            'product_url': it.get('product_link', ''),
            'shop_id': shop_id, 'item_id': item_id,
            'has_video': bool((info.get('video_info_list') or [])),
        })
    return items

async def chk_video(page, sid, iid):
    if not sid or not iid: return False
    js = """
    (async () => {
      try {
        const r = await fetch('https://shopee.tw/api/v4/item/get?itemid=IID&shopid=SID', {credentials:'include'});
        const d = await r.json();
        return (d && d.data && d.data.video_info_list && d.data.video_info_list.length > 0);
      } catch(e) { return false; }
    })()
    """.replace('IID', str(iid)).replace('SID', str(sid))
    return bool(await safe_eval(page, js))

async def check_reviews(page, sid, iid, debug=False):
    """回傳 (has_recent_review, video_review_count)"""
    if not sid or not iid: return False, 0
    js = """
    (async () => {
      try {
        // 嘗試 v4/product/get_ratings
        const r = await fetch('https://shopee.tw/api/v4/product/get_ratings?itemid=IID&shopid=SID&offset=0&limit=50', {credentials:'include'});
        const d = await r.json();
        if (!d || !d.data || !d.data.ratings) {
          // error_not_found = API 被 block，回傳 -1 表示跳過
          const isBlocked = d && d.error;
          return {recent: false, vid_count: isBlocked ? -1 : 0, raw: JSON.stringify(d).substring(0, 200)};
        }
        const now = Math.floor(Date.now() / 1000);
        const cutoff = now - DAYS * 86400;
        let recent = false, vid_count = 0;
        for (const rv of d.data.ratings) {
          if (rv.ctime && rv.ctime >= cutoff) recent = true;
          if ((rv.videos && rv.videos.length > 0) || rv.video_url) vid_count++;
        }
        return {recent, vid_count, total: d.data.ratings.length};
      } catch(e) { return {recent: false, vid_count: 0, err: e.toString()}; }
    })()
    """.replace('IID', str(iid)).replace('SID', str(sid)).replace('DAYS', str(REVIEW_DAYS))
    res = await safe_eval(page, js) or {}
    if debug: print(f'\n    [review debug] sid={sid} iid={iid} -> {res}')
    return bool(res.get('recent')), int(res.get('vid_count', 0))

async def get_link(page, url):
    if not url: return ''
    enc = quote(url, safe='')
    js = """
    (async () => {
      try {
        const r = await fetch('https://affiliate.shopee.tw/api/v3/product/get_affiliate_link?product_url=URL', {credentials:'include'});
        const d = await r.json();
        return (d && d.data && (d.data.short_link || d.data.link)) || '';
      } catch(e) { return ''; }
    })()
    """.replace('URL', enc)
    return await safe_eval(page, js) or ''

def build_excel(products, append=False):
    thin = Side(style='thin', color='DDDDDD')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdrs = ['編號','品名','分潤連結','價格','分潤率','銷量','對應關鍵字','文案','標題','狀態']
    wds  = [8, 55, 48, 10, 10, 10, 16, 30, 30, 12]

    if append and os.path.exists(OUTPUT):
        wb = openpyxl.load_workbook(OUTPUT)
        ws = wb.active
        start_row = ws.max_row + 1
        start_no  = ws.max_row      # 編號接續（第1行是header）
        print(f'  Append 模式：從第 {start_row} 行接續（編號從 {start_no} 開始）')
    else:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '蝦皮關鍵字選品'
        for ci, (h, w) in enumerate(zip(hdrs, wds), 1):
            c = ws.cell(1, ci, h)
            c.font = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=11)
            c.fill = PatternFill('solid', fgColor='C0392B')
            c.border = bd
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 28
        start_row = 2; start_no = 1

    for i, p in enumerate(products):
        ri = start_row + i
        no = start_no + i
        fill = PatternFill('solid', fgColor='FFF5F5' if ri % 2 == 0 else 'FFFFFF')
        vals = [no, p.get('name',''), p.get('affiliate_link',''), f"${p.get('price','0')}",
                p.get('comm_rate',''), p.get('sales',0), p.get('keyword',''), '', '', '']
        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci, val); c.fill = fill; c.border = bd
            c.font = Font(name='Arial', size=10, color='0563C1', underline='single') if ci == 3 \
                     else Font(name='微軟正黑體', size=10)
            c.alignment = Alignment(horizontal='left' if ci in (2, 3) else 'center',
                                    vertical='center', wrap_text=(ci == 2))
        ws.row_dimensions[ri].height = 32

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:J{ws.max_row}'
    wb.save(OUTPUT)
    print(f'Excel 已儲存：{OUTPUT}（共 {ws.max_row - 1} 筆）')

async def main():
    print('=' * 50)
    print('蝦皮關鍵字選品 Skill B - 2026年3-4月')
    print('=' * 50)

    async with async_playwright() as p:
        print('\n連接 Chrome (port 9222)...')
        browser = await p.chromium.connect_over_cdp('http://127.0.0.1:9222')
        ctx = browser.contexts[0]

        # 找 affiliate 頁面
        pg_aff = next((pg for pg in ctx.pages if 'affiliate.shopee.tw' in pg.url), None)
        if not pg_aff:
            pg_aff = await ctx.new_page()
            await pg_aff.goto('https://affiliate.shopee.tw/offer/product_offer',
                              wait_until='domcontentloaded', timeout=20000)
            await asyncio.sleep(3)
        print(f'affiliate 頁面：{pg_aff.url}')

        # 找或開 shopee.tw 頁面（用來查影片/評論）
        pg_main = next((pg for pg in ctx.pages
                        if pg.url.startswith('https://shopee.tw') and 'affiliate' not in pg.url), None)
        if not pg_main:
            pg_main = await ctx.new_page()
        # 若停在驗證/錯誤頁，重新導向首頁
        if any(x in pg_main.url for x in ['captcha', 'verify', 'traffic']):
            print('  shopee.tw 驗證頁，重新導向首頁...')
            await pg_main.goto('https://shopee.tw', wait_until='domcontentloaded', timeout=20000)
            await asyncio.sleep(3)
        print(f'shopee 頁面：{pg_main.url[:60]}')

        # API 測試
        test = await search_kw(pg_aff, '手機殼', limit=2)
        print(f'API 測試：{"OK - " + str(len(test)) + "筆" if test else "FAIL"}')

        # 載入歷史商品 ID（跨次去重用）
        prev_uids = set(load_product_history().keys())
        print(f'  歷史商品 {len(prev_uids)} 筆（本次將排除）')

        # [2] 搜尋
        print(f'\n[2/4] 搜尋關鍵字（目標 {BATCH} 筆，每關鍵字上限 {MAX_PER_KW} 筆）...')
        all_p = []; seen = set(); kw_count = {}
        for kw in KEYWORDS:
            if len(all_p) >= BATCH: break
            items = await search_kw(pg_aff, kw, limit=40)
            added = 0
            for item in items:
                if kw_count.get(kw, 0) >= MAX_PER_KW: break  # 單一關鍵字上限
                uid = f"{item['shop_id']}_{item['item_id']}"
                if uid in seen or uid == '_': continue
                if uid in prev_uids: continue  # 已在歷史中，跳過
                seen.add(uid); all_p.append(item)
                kw_count[kw] = kw_count.get(kw, 0) + 1
                added += 1
            print(f'  [{kw}] {added}筆 累計{len(all_p)}')
            await asyncio.sleep(random.uniform(0.4, 0.8))
        print(f'共收集 {len(all_p)} 筆（已排除歷史重複）')

        # [3] 過濾：有分潤連結即可（影片由 video_maker 判斷）
        print(f'\n[3/4] 過濾（有分潤連結即可）...')
        valid = []; nl = 0
        for i, item in enumerate(all_p):
            if len(valid) >= TARGET: break
            print(f'  [{i+1:3}/{len(all_p)}] {item["name"][:20]}...', end=' ', flush=True)

            # 只檢查分潤連結
            lnk = item.get('affiliate_link', '')
            if not lnk:
                lnk = await get_link(pg_aff, item['product_url'])
                item['affiliate_link'] = lnk
            if not lnk:
                print('-> 無連結'); nl += 1; continue

            print(f'✅ ({len(valid)+1}/{TARGET})')
            valid.append(item)
            await asyncio.sleep(0.1)
        print(f'\n有效 {len(valid)} | 無連結 {nl}')

        pass  # browser stays open

    # [4] Excel
    print('\n[4/4] 輸出 Excel...')
    final = valid[:TARGET]
    if len(final) == 0:
        print('⚠️  0 筆有效商品（可能是 shopee.tw 驗證碼擋住），不覆寫 Excel。')
        return
    append_mode = os.getenv('KEYWORD_APPEND', '0') == '1'
    build_excel(final, append=append_mode)

    # 記錄本次選到的商品（跨次去重）
    save_product_history(final)
    print('\n完成！')

asyncio.run(main())
